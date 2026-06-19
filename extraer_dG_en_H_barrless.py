#!/usr/bin/env python3
"""
deltag_barrless.py
Calcula DeltaGr(T) en hartrees para las reacciones sin barrera del pif.chem.
Ejecutar desde: pilgrim_C6H5/
"""
import os
import re

# ============================================================
# CONFIGURACIÓN
# ============================================================

CHEM_FILE   = "pif.chem"
PLG_DIR     = "3-PLG_OUTPUT"
OUTPUT_XLSX = "deltaG_barrless.xlsx"

# ============================================================
# PARSEAR pif.chem — solo líneas barrierless (etiqueta con letra)
# ============================================================

def parse_barrierless(chem_file):
    reactions = []
    pattern = re.compile(
        r"^([a-z][a-z0-9]*)\s*:\s*(\S+)\s*-->\s*\S+\s*-->\s*(\S+)\s*\+\s*(\S+)"
    )
    with open(chem_file) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            m = pattern.match(line)
            if m:
                reactions.append({
                    "label"   : m.group(1),
                    "reactivo": m.group(2),
                    "prod_a"  : m.group(3),
                    "prod_b"  : m.group(4),
                })
    return reactions

# ============================================================
# EXTRAER G(T) de la columna v = 1 cm^3
# ============================================================

def extract_gibbs(species, plg_dir):
    fname = os.path.join(plg_dir, f"pfn.{species}.slevel.txt")
    if not os.path.isfile(fname):
        raise FileNotFoundError(f"No encontrado: {fname}")

    gibbs = {}
    in_section = False

    with open(fname) as f:
        for line in f:
            if "Gibbs free energy (hartree):" in line:
                in_section = True
                continue
            if not in_section:
                continue
            if in_section and "---" in line and gibbs:
                break
            # Columnas: T | v=1cm^3 | v=kbT/p0
            m = re.search(r"\|\s*([\d.]+)\s*\|\s*([-\d.]+)\s*\|\s*([-\d.]+)", line)
            if m:
                T = float(m.group(1))
                G = float(m.group(2))   # v = 1 cm^3
                gibbs[T] = G

    if not gibbs:
        raise ValueError(f"No se encontraron datos Gibbs en: {fname}")
    return gibbs

# ============================================================
# CALCULAR DeltaGr(T) = G(prod_a) + G(prod_b) - G(reactivo)
# ============================================================

def calc_delta_g(reactivo, prod_a, prod_b, plg_dir):
    G_r  = extract_gibbs(reactivo, plg_dir)
    G_pa = extract_gibbs(prod_a,   plg_dir)
    G_pb = extract_gibbs(prod_b,   plg_dir)

    temps = sorted(set(G_r) & set(G_pa) & set(G_pb))
    results = {
        T: {
            "G_reactivo" : G_r[T],
            "G_producto" : G_pa[T] + G_pb[T],
            "DeltaG"     : G_pa[T] + G_pb[T] - G_r[T],
        }
        for T in temps
    }
    return results

# ============================================================
# MAIN
# ============================================================

def main():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    reactions = parse_barrierless(CHEM_FILE)
    if not reactions:
        print("No se encontraron reacciones barrierless en pif.chem.")
        return

    total = len(reactions)
    ok    = 0
    error = 0

    wb = Workbook()
    ws = wb.active
    ws.title = "DeltaG_barrless"

    current_row = 1

    for rxn in reactions:
        label    = rxn["label"]
        reactivo = rxn["reactivo"]
        prod_a   = rxn["prod_a"]
        prod_b   = rxn["prod_b"]

        # ── pantalla ──
        print(f"\n{'='*70}")
        print(f"  Reaccion {label}: {reactivo} --> {prod_a} + {prod_b}")
        print(f"{'='*70}")

        try:
            results = calc_delta_g(reactivo, prod_a, prod_b, PLG_DIR)
            ok += 1
        except (FileNotFoundError, ValueError) as e:
            print(f"  ERROR: {e}")
            error += 1
            # Escribir error en xlsx
            ws.cell(row=current_row, column=1,
                    value=f"Reaccion {label}: {reactivo} --> {prod_a} + {prod_b}").font = Font(bold=True)
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"ERROR: {e}")
            current_row += 3
            continue

        # ── pantalla: cabecera ──
        print(f"  {'T (K)':<10} {'G_reactivo (H)':<22} {'G_producto (H)':<22} {'DeltaG (H)':<22}")
        print(f"  {'-'*76}")

        # ── pantalla: datos ──
        for T in sorted(results):
            r = results[T]
            print(f"  {T:<10.2f} {r['G_reactivo']:>+22.8f} {r['G_producto']:>+22.8f} {r['DeltaG']:>+22.8f}")

        # ── xlsx: cabecera de reaccion ──
        ws.cell(row=current_row, column=1,
                value=f"Reaccion {label}: {reactivo} --> {prod_a} + {prod_b}").font = Font(bold=True)
        current_row += 1

        # ── xlsx: cabecera de columnas ──
        headers = ["T (K)", "G_reactivo (hartree)", "G_producto (hartree)", "DeltaG (hartree)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=h).font = Font(bold=True)
        current_row += 1

        # ── xlsx: datos ──
        for T in sorted(results):
            r = results[T]
            ws.cell(row=current_row, column=1, value=T)
            ws.cell(row=current_row, column=2, value=round(r["G_reactivo"], 8))
            ws.cell(row=current_row, column=3, value=round(r["G_producto"], 8))
            ws.cell(row=current_row, column=4, value=round(r["DeltaG"],     8))
            current_row += 1

        # Dos filas vacías entre reacciones
        current_row += 2

    # ── resumen pantalla ──
    resumen = (
        f"\n{'='*70}\n"
        f"  Reacciones totales en pif.chem : {total}\n"
        f"  Procesadas correctamente       : {ok}\n"
        f"  Con error                      : {error}\n"
        f"{'='*70}\n"
    )
    print(resumen)

    wb.save(OUTPUT_XLSX)
    print(f"Archivo guardado: {OUTPUT_XLSX}")
    print("Proceso completado correctamente ✔\n")

if __name__ == "__main__":
    main()
